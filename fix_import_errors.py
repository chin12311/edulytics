import openpyxl
import re

def fix_excel_file(input_file, output_file):
    """Fix role capitalization and username issues in Excel file."""
    
    # Load the workbook
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    # Get header row to find column indices
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[cell.value.lower().strip()] = col_idx
    
    print(f"Found headers: {list(headers.keys())}")
    
    # Get column indices
    role_col = headers.get('role')
    username_col = headers.get('username')
    
    if not role_col:
        print("ERROR: 'role' column not found!")
        return
    
    fixed_rows = 0
    fixed_usernames = 0
    
    # Process each row (starting from row 2, skipping header)
    for row_idx in range(2, ws.max_row + 1):
        row_changed = False
        
        # Fix role column
        role_cell = ws.cell(row=row_idx, column=role_col)
        if role_cell.value:
            original_role = str(role_cell.value).strip()
            # Capitalize first letter, lowercase rest
            fixed_role = original_role.capitalize()
            
            # Map to valid roles
            role_mapping = {
                'student': 'Student',
                'faculty': 'Faculty',
                'dean': 'Dean',
                'coordinator': 'Coordinator',
                'admin': 'Admin'
            }
            
            if original_role.lower() in role_mapping:
                new_role = role_mapping[original_role.lower()]
                if role_cell.value != new_role:
                    role_cell.value = new_role
                    row_changed = True
                    print(f"Row {row_idx}: Fixed role '{original_role}' -> '{new_role}'")
        
        # Fix username column (remove invalid characters)
        if username_col:
            username_cell = ws.cell(row=row_idx, column=username_col)
            if username_cell.value:
                original_username = str(username_cell.value).strip()
                # Keep only letters, numbers, dots, hyphens, and underscores
                fixed_username = re.sub(r'[^a-zA-Z0-9._-]', '', original_username)
                
                if original_username != fixed_username:
                    username_cell.value = fixed_username
                    fixed_usernames += 1
                    print(f"Row {row_idx}: Fixed username '{original_username}' -> '{fixed_username}'")
                    row_changed = True
        
        if row_changed:
            fixed_rows += 1
    
    # Save the fixed workbook
    wb.save(output_file)
    print(f"\n✅ Fixed {fixed_rows} rows")
    print(f"✅ Fixed {fixed_usernames} usernames")
    print(f"✅ Saved to: {output_file}")

if __name__ == "__main__":
    # Update these paths to match your files
    input_file = "accounts_ready.xlsx"
    output_file = "accounts_fixed.xlsx"
    
    print(f"Processing {input_file}...")
    fix_excel_file(input_file, output_file)
    print("\nDone! You can now import accounts_fixed.xlsx")
