"""
Service for importing and exporting accounts to/from Excel files.
"""
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.contrib.auth.models import User
from django.db import transaction
from main.models import UserProfile, Section, Role
from django.core.exceptions import ValidationError
from main.validation_utils import AccountValidator


class AccountImportExportService:
    """Handle import and export of accounts to Excel files."""
    
    # Define which fields we need from the Excel file
    REQUIRED_FIELDS = ['username', 'email', 'password', 'display_name', 'role']
    STUDENT_FIELDS = ['student_number', 'course', 'section']
    STAFF_FIELDS = ['institute']
    
    @staticmethod
    def export_accounts_to_excel():
        """
        Export all accounts from the system to an Excel file.
        Returns: BytesIO object with the Excel file.
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Accounts"
        
        # Define header style
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create headers
        headers = [
            'Username',
            'Email',
            'Display Name',
            'Role',
            'Student Number',
            'Course',
            'Section',
            'Institute',
            'Date Joined'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Set column widths
        column_widths = [15, 20, 20, 15, 15, 30, 15, 20, 15]
        for col, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[chr(64 + col)].width = width
        
        # Add user data - sorted by role
        users = User.objects.select_related('userprofile').all()
        
        # Filter out users without userprofile
        users_with_profile = [u for u in users if hasattr(u, 'userprofile') and u.userprofile is not None]
        
        # Sort users by role priority: Student, Faculty, Coordinator, Dean, Admin
        role_order = {
            Role.STUDENT: 0,
            Role.FACULTY: 1,
            Role.COORDINATOR: 2,
            Role.DEAN: 3,
            Role.ADMIN: 4
        }
        
        # Define role colors for visual grouping
        role_colors = {
            Role.STUDENT: "E8F5E9",      # Light green
            Role.FACULTY: "E3F2FD",      # Light blue
            Role.COORDINATOR: "FFF3E0",  # Light orange
            Role.DEAN: "F3E5F5",         # Light purple
            Role.ADMIN: "FCE4EC"         # Light pink
        }
        
        sorted_users = sorted(users_with_profile, key=lambda u: (role_order.get(u.userprofile.role, 5), u.username))
        
        current_role = None
        for row, user in enumerate(sorted_users, 2):
            profile = user.userprofile
            
            # Add a section separator if role changes
            if current_role != profile.role and current_role is not None:
                # Add an empty row with role separator
                row += 1
            
            current_role = profile.role
            
            row_data = [
                user.username,
                user.email,
                profile.display_name or '',
                profile.get_role_display(),
                profile.studentnumber if profile.role == Role.STUDENT else '',
                profile.course if profile.role == Role.STUDENT else '',
                str(profile.section) if profile.role == Role.STUDENT and profile.section else '',
                profile.institute if profile.role in [Role.DEAN, Role.FACULTY, Role.COORDINATOR] else '',
                user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else ''
            ]
            
            # Get the background color for this role
            bg_color = role_colors.get(profile.role, "FFFFFF")
            role_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            
            for col, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border
                cell.fill = role_fill  # Apply role-based color
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    @staticmethod
    def import_accounts_from_excel(excel_file):
        """
        Import accounts from an Excel file with comprehensive validation.
        
        Args:
            excel_file: The uploaded Excel file
            
        Returns:
            dict with 'success' (bool), 'created' (int), 'updated' (int), 'errors' (list)
        """
        result = {
            'success': False,
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': []
        }
        
        try:
            # Load the workbook
            workbook = load_workbook(excel_file)
            worksheet = workbook.active
            
            # Get headers from first row
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    header = str(cell.value).strip().lower().replace(' ', '_')
                    headers.append(header)
            
            # Validate headers
            required = {'username', 'email', 'password', 'display_name', 'role'}
            if not required.issubset(set(headers)):
                missing = required - set(headers)
                result['errors'].append(f"Missing required columns: {', '.join(missing)}")
                return result
            
            # Create a mapping of header name to column index
            header_mapping = {header: idx + 1 for idx, header in enumerate(headers)}
            
            # Track duplicate checking
            usernames_in_batch = set()
            emails_in_batch = set()
            
            # Process each row
            with transaction.atomic():
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), 2):
                    try:
                        # Extract values
                        row_data = {}
                        for header, col_idx in header_mapping.items():
                            cell = row[col_idx - 1]
                            row_data[header] = cell.value if cell.value else ''
                        
                        # Get required fields
                        username = str(row_data.get('username', '')).strip()
                        email = str(row_data.get('email', '')).strip()
                        password = str(row_data.get('password', '')).strip()
                        display_name = str(row_data.get('display_name', '')).strip()
                        role = str(row_data.get('role', '')).strip()
                        
                        # Validate required fields
                        if not all([username, email, password, display_name, role]):
                            result['errors'].append(f"Row {row_idx}: Missing required fields")
                            result['skipped'] += 1
                            continue
                        
                        # ✅ VALIDATE USERNAME using AccountValidator
                        is_update = User.objects.filter(username=username).exists()
                        exclude_id = User.objects.filter(username=username).values_list('id', flat=True).first() if is_update else None
                        
                        valid, msg = AccountValidator.validate_username(
                            username,
                            exclude_user_id=exclude_id
                        )
                        if not valid:
                            result['errors'].append(f"Row {row_idx}: {msg}")
                            result['skipped'] += 1
                            continue
                        
                        # Check for duplicates within batch
                        if username in usernames_in_batch and not is_update:
                            result['errors'].append(f"Row {row_idx}: Duplicate username '{username}' in batch")
                            result['skipped'] += 1
                            continue
                        
                        # ✅ VALIDATE EMAIL using AccountValidator
                        valid, msg = AccountValidator.validate_email(
                            email,
                            exclude_user_id=exclude_id
                        )
                        if not valid:
                            result['errors'].append(f"Row {row_idx}: {msg}")
                            result['skipped'] += 1
                            continue
                        
                        # Check for email duplicates within batch
                        if email in emails_in_batch and not is_update:
                            result['errors'].append(f"Row {row_idx}: Duplicate email '{email}' in batch")
                            result['skipped'] += 1
                            continue
                        
                        # ✅ VALIDATE PASSWORD (only for new users or if provided)
                        if not is_update or (password and password != '***EXISTING***'):
                            valid, msg = AccountValidator.validate_password(password)
                            if not valid:
                                result['errors'].append(f"Row {row_idx}: Password invalid - {msg}")
                                result['skipped'] += 1
                                continue
                        
                        # ✅ VALIDATE DISPLAY NAME using AccountValidator
                        valid, msg = AccountValidator.validate_display_name(display_name)
                        if not valid:
                            result['errors'].append(f"Row {row_idx}: {msg}")
                            result['skipped'] += 1
                            continue
                        
                        # Validate role
                        valid, msg = AccountValidator.validate_role(role)
                        if not valid:
                            result['errors'].append(f"Row {row_idx}: {msg}")
                            result['skipped'] += 1
                            continue
                        
                        # Role is already in correct format (Student, Faculty, etc.)
                        role_enum = role
                        
                        # ✅ VALIDATE ROLE-SPECIFIC FIELDS
                        if role_enum == Role.STUDENT:
                            student_number = str(row_data.get('student_number', '')).strip()
                            course = str(row_data.get('course', '')).strip()
                            section_value = str(row_data.get('section', '')).strip()
                            
                            # Validate student number
                            if not student_number:
                                result['errors'].append(f"Row {row_idx}: Student number required for students")
                                result['skipped'] += 1
                                continue
                            
                            valid, msg = AccountValidator.validate_student_number(student_number)
                            if not valid:
                                result['errors'].append(f"Row {row_idx}: {msg}")
                                result['skipped'] += 1
                                continue
                            
                            # Validate course
                            valid, msg = AccountValidator.validate_course(course)
                            if not valid:
                                result['errors'].append(f"Row {row_idx}: {msg}")
                                result['skipped'] += 1
                                continue
                            
                            # Validate section if provided (supports "CODE (Label)" format)
                            section_obj = None
                            if section_value:
                                normalized_section = section_value.split('(')[0].strip()
                                if normalized_section:
                                    section_obj = Section.objects.filter(code=normalized_section).first()
                                    if not section_obj:
                                        try:
                                            section_obj = Section.objects.get(id=int(normalized_section))
                                        except (ValueError, Section.DoesNotExist):
                                            section_obj = None
                                if not section_obj:
                                    result['errors'].append(f"Row {row_idx}: Section '{section_value}' not found")
                                    result['skipped'] += 1
                                    continue
                        
                        elif role_enum in [Role.DEAN, Role.FACULTY, Role.COORDINATOR]:
                            institute = str(row_data.get('institute', '')).strip()
                            
                            # Validate institute
                            valid, msg = AccountValidator.validate_institute(institute)
                            if not valid:
                                result['errors'].append(f"Row {row_idx}: {msg}")
                                result['skipped'] += 1
                                continue
                        
                        # Check if user exists
                        user_exists = User.objects.filter(username=username).exists()
                        
                        if user_exists:
                            # Update existing user
                            user = User.objects.get(username=username)
                            user.email = email
                            if password and password != '***EXISTING***':  # Allow skipping password update
                                user.set_password(password)
                            user.save()
                            
                            profile = user.userprofile
                            profile.display_name = display_name
                            profile.role = role_enum
                            
                            # Handle student-specific fields
                            if role_enum == Role.STUDENT:
                                profile.studentnumber = str(row_data.get('student_number', '')).strip()
                                profile.course = str(row_data.get('course', '')).strip()
                                
                                # Handle section
                                section_value = str(row_data.get('section', '')).strip()
                                if section_value:
                                    normalized_section = section_value.split('(')[0].strip()
                                    section = Section.objects.filter(code=normalized_section).first()
                                    if not section:
                                        try:
                                            section = Section.objects.get(id=int(normalized_section))
                                        except (ValueError, Section.DoesNotExist):
                                            section = None
                                    if section:
                                        profile.section = section
                                    else:
                                        result['errors'].append(f"Row {row_idx}: Section '{section_value}' not found")
                            
                            # Handle staff-specific fields
                            elif role_enum in [Role.DEAN, Role.FACULTY, Role.COORDINATOR]:
                                profile.institute = str(row_data.get('institute', '')).strip()
                            
                            profile.save()
                            result['updated'] += 1
                        else:
                            # Create new user
                            user = User.objects.create_user(
                                username=username,
                                email=email,
                                password=password
                            )
                            
                            profile = user.userprofile
                            profile.display_name = display_name
                            profile.role = role_enum
                            
                            # Handle student-specific fields
                            if role_enum == Role.STUDENT:
                                profile.studentnumber = str(row_data.get('student_number', '')).strip()
                                profile.course = str(row_data.get('course', '')).strip()
                                
                                # Handle section
                                section_value = str(row_data.get('section', '')).strip()
                                if section_value:
                                    normalized_section = section_value.split('(')[0].strip()
                                    section = Section.objects.filter(code=normalized_section).first()
                                    if not section:
                                        try:
                                            section = Section.objects.get(id=int(normalized_section))
                                        except (ValueError, Section.DoesNotExist):
                                            section = None
                                    if section:
                                        profile.section = section
                                    else:
                                        result['errors'].append(f"Row {row_idx}: Section '{section_value}' not found")
                            
                            # Handle staff-specific fields
                            elif role_enum in [Role.DEAN, Role.FACULTY, Role.COORDINATOR]:
                                profile.institute = str(row_data.get('institute', '')).strip()
                            
                            profile.save()
                            result['created'] += 1
                            
                            # Track batch duplicates
                            usernames_in_batch.add(username)
                            emails_in_batch.add(email)
                    
                    except Exception as e:
                        result['errors'].append(f"Row {row_idx}: {str(e)}")
                        result['skipped'] += 1
            
            result['success'] = True
            
        except Exception as e:
            result['errors'].append(f"Failed to read Excel file: {str(e)}")
        
        return result
