import openpyxl
import re

input_file = r"C:\Users\ADMIN\Downloads\accounts_export.xlsx"
output_file = r"C:\Users\ADMIN\eval\evaluation\accounts_final.xlsx"

# Load workbook
wb = openpyxl.load_workbook(input_file)
ws = wb.active

# Get headers
headers = [cell.value for cell in ws[1]]
print(f"Original headers: {headers}")

# Find column indices
username_idx = headers.index('Username') + 1 if 'Username' in headers else 1
email_idx = headers.index('Email') + 1 if 'Email' in headers else 2
role_idx = headers.index('Role') + 1 if 'Role' in headers else 4

# Check if password and display_name exist
has_password = 'password' in [h.lower() if h else '' for h in headers]
has_display_name = any('display' in str(h).lower() if h else '' for h in headers)

if not has_password:
    # Insert password column after email
    ws.insert_cols(email_idx + 1)
    ws.cell(1, email_idx + 1, 'password')
    print("Added 'password' column")
    
    # Update display_name index if it exists
    if has_display_name:
        display_idx = email_idx + 2
    else:
        display_idx = email_idx + 1
else:
    password_idx = next(i+1 for i, h in enumerate(headers) if h and 'password' in h.lower())
    display_idx = password_idx + 1

if not has_display_name:
    # Insert display_name column
    ws.insert_cols(display_idx + 1)
    ws.cell(1, display_idx + 1, 'display_name')
    print("Added 'display_name' column")

# Get final header positions after insertions
headers_final = [cell.value for cell in ws[1]]
username_idx = headers_final.index('Username') + 1
email_idx = headers_final.index('Email') + 1
password_idx = next(i+1 for i, h in enumerate(headers_final) if h and 'password' in h.lower())
display_idx = next(i+1 for i, h in enumerate(headers_final) if h and 'display' in h.lower())
role_idx = headers_final.index('Role') + 1

print(f"\nColumn indices: username={username_idx}, email={email_idx}, password={password_idx}, display_name={display_idx}, role={role_idx}")

# Role mapping (normalize variations)
role_map = {
    'student': 'Student',
    'faculty': 'Faculty', 
    'teacher': 'Faculty',
    'dean': 'Dean',
    'coordinator': 'Coordinator',
    'admin': 'Admin',
    'administrator': 'Admin'
}

# Clean display names - only letters, spaces, hyphens, apostrophes
def clean_display_name(name):
    if not name:
        return ''
    # Remove numbers and special chars except space, hyphen, apostrophe
    cleaned = re.sub(r'[^a-zA-Z\s\-\']', '', str(name))
    # Remove extra spaces
    cleaned = ' '.join(cleaned.split())
    return cleaned

# Process rows
fixed = 0
for row_idx in range(2, ws.max_row + 1):
    username = ws.cell(row_idx, username_idx).value
    
    # Skip empty rows
    if not username:
        continue
    
    # Fix password
    ws.cell(row_idx, password_idx, 'Password123!')
    
    # Fix display_name
    current_display = ws.cell(row_idx, display_idx).value
    if current_display:
        cleaned = clean_display_name(current_display)
        ws.cell(row_idx, display_idx, cleaned)
    else:
        # Generate from username
        cleaned = clean_display_name(str(username).replace('_', ' ').replace('.', ' '))
        cleaned = ' '.join(word.capitalize() for word in cleaned.split())
        ws.cell(row_idx, display_idx, cleaned)
    
    # Fix role - normalize to proper case
    current_role = ws.cell(row_idx, role_idx).value
    if current_role:
        role_lower = str(current_role).strip().lower()
        proper_role = role_map.get(role_lower, current_role)
        ws.cell(row_idx, role_idx, proper_role)
        
    fixed += 1

# Save
wb.save(output_file)
print(f"\n✓ Fixed {fixed} accounts")
print(f"✓ Saved to: {output_file}")
print(f"\nReady to import!")
