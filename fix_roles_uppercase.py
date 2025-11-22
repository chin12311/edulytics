import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\ADMIN\eval\evaluation\accounts_final.xlsx')
ws = wb.active

headers = [c.value for c in ws[1]]
role_idx = headers.index('role') + 1

print('Converting roles to UPPERCASE...')
count = 0
for r in range(2, ws.max_row + 1):
    cell = ws.cell(r, role_idx)
    if cell.value:
        cell.value = str(cell.value).upper()
        count += 1

wb.save(r'C:\Users\ADMIN\eval\evaluation\accounts_final.xlsx')
print(f'✓ Fixed {count} roles to UPPERCASE')
print('✓ File ready for import!')
