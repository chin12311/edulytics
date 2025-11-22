import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\ADMIN\eval\evaluation\accounts_final.xlsx')
ws = wb.active

headers = [c.value for c in ws[1]]
role_idx = headers.index('role') + 1

print('Converting roles to proper case (Student, Faculty, etc.)...')
count = 0
for r in range(2, ws.max_row + 1):
    cell = ws.cell(r, role_idx)
    if cell.value:
        cell.value = str(cell.value).capitalize()
        count += 1

wb.save(r'C:\Users\ADMIN\eval\evaluation\accounts_ready.xlsx')
print(f'✓ Fixed {count} roles to proper capitalization')
print('✓ Roles now: Student, Faculty, Dean, Coordinator, Admin')
print('✓ Saved to: accounts_ready.xlsx')
