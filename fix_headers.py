import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\ADMIN\eval\evaluation\accounts_final.xlsx')
ws = wb.active

# Required header mapping
header_map = {
    'Username': 'username',
    'Email': 'email', 
    'password': 'password',
    'display_name': 'display_name',
    'Role': 'role'
}

headers = [c.value for c in ws[1]]
print('Before:', headers[:6])

for i, h in enumerate(headers, 1):
    if h in header_map:
        ws.cell(1, i, header_map[h])
        
new_headers = [c.value for c in ws[1]]
print('After:', new_headers[:6])

wb.save(r'C:\Users\ADMIN\eval\evaluation\accounts_final.xlsx')
print('✓ Headers normalized to lowercase')
