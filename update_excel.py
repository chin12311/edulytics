import openpyxl
import sys

# Get the input file path from user  
input_file = r"C:\Users\ADMIN\Downloads\accounts_export.xlsx"

try:
    # Load the workbook
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # Get current headers
    headers = [cell.value for cell in ws[1]]
    print(f'Current headers: {headers}')

    # Check if password and display_name already exist
    if 'password' in headers and 'display_name' in headers:
        print('Columns already exist!')
        sys.exit(0)

    # Find email column index
    if 'email' in headers:
        email_col_idx = headers.index('email') + 1
    else:
        email_col_idx = 3  # Default to column C if email not found

    # Insert two new columns after email
    ws.insert_cols(email_col_idx + 1, 2)

    # Set new headers
    ws.cell(1, email_col_idx + 1, 'password')
    ws.cell(1, email_col_idx + 2, 'display_name')

    # Fill data for each row
    for row_idx in range(2, ws.max_row + 1):
        # Set default password with special character
        ws.cell(row_idx, email_col_idx + 1, 'Password123!')
        
        # Set display_name from username (column A)
        username_cell = ws.cell(row_idx, 1)
        if username_cell.value:
            username = str(username_cell.value)
            # Convert username to display name
            display_name = ' '.join(word.capitalize() for word in username.replace('_', ' ').replace('.', ' ').split())
            ws.cell(row_idx, email_col_idx + 2, display_name)

    # Save the updated file
    output_path = r'C:\Users\ADMIN\eval\evaluation\accounts_updated.xlsx'
    wb.save(output_path)
    
    print(f'\n✓ Success!')
    print(f'✓ Updated file saved to: {output_path}')
    print(f'✓ Total accounts: {ws.max_row - 1}')
    print(f'✓ Added columns: password (default: Password123), display_name')

except FileNotFoundError:
    print(f'Error: File not found at {input_file}')
    print('Please check the file path and try again.')
except Exception as e:
    print(f'Error: {e}')
