import openpyxl

def inspect_excel_file(filename):
    """Inspect the Excel file to see actual role values."""
    
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Get header row
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[cell.value.lower().strip()] = col_idx
    
    role_col = headers.get('role')
    
    print(f"Inspecting roles in {filename}:")
    print(f"Headers: {list(headers.keys())}\n")
    
    # Check first 10 data rows
    print("First 10 data rows:")
    for row_idx in range(2, min(12, ws.max_row + 1)):
        role_cell = ws.cell(row=row_idx, column=role_col)
        username_cell = ws.cell(row=row_idx, column=headers.get('username'))
        print(f"Row {row_idx}: username='{username_cell.value}', role='{role_cell.value}' (type: {type(role_cell.value).__name__})")

if __name__ == "__main__":
    inspect_excel_file("accounts_ready.xlsx")
